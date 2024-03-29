from flask import Flask, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook
import jwt
import datetime

app = Flask(__name__)
app.config["SECRET_KEY"] = "wshuduihfewih7rh43y128qhed73259gjp7FL"  # Change this to a secure secret key
app.config["SQLALCHEMY_DATABASE_URI"] = 'sqlite:///db.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
ma = Marshmallow(app)

class Fparser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    city = db.Column(db.String(200), nullable=False)

class parserschma(ma.Schema):
    class Meta:
        fields = ['id', 'name', 'age', 'city']

parser_schema = parserschma()
parser_schemas = parserschma(many=True)

# Function to generate JWT token
def generate_token(username):
    # header = {"algorithm": "HS256"}
    payload = {
        'username': username,
        'exp': (datetime.datetime.utcnow()) + datetime.timedelta(days=1)  # Token expiration time
    }
    token = jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')
    return token

# Function to verify JWT token
def verify_token(func):
    def wrapper(*args, **kwargs):
        token = request.headers.get('Authorization')
        print(token)
        if not token:
            return jsonify({"msg": "Missing token"}), 401

        # try:
        #     payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        #     request.username = payload['username']
        #     return func(*args, **kwargs)
        # except jwt.ExpiredSignatureError:
        #     return jsonify({"msg": "Token expired. Please log in again."}), 401
        # except jwt.InvalidTokenError:
        #     return jsonify({"msg": "Invalid token. Please log in again."}), 401
        try:
    # Remove padding characters from the token
            # token = token.remove('Bear/e')
            token = (token.replace('Bearer ', ''))
    # Decode the JWT token
            payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            request.username = payload['username']
            return func(*args, **kwargs)
        except jwt.ExpiredSignatureError:
            return jsonify({"msg": "Token expired. Please log in again."}), 401
    # Preserve the original function name
    wrapper.__name__ = func.__name__
    return wrapper



# Authentication route
@app.route('/login', methods=['POST'])
def login():
    username = request.json.get('username', None)
    password = request.json.get('password', None)

    # Add your user authentication logic here (e.g., check username and password against the database)
    # For simplicity, let's assume you have a hardcoded user
    if username == 'user' and password == 'password':
        access_token = generate_token(username)
        return jsonify(access_token=access_token), 200
    else:
        return jsonify({"msg": "Bad username or password"}), 401

# CRUD operations with JWT authentication
@app.route('/add', methods=['POST'])
def add_data():
    if request.method == 'POST':
        excel_data = request.files['Schedule']
        Mydata = load_workbook(excel_data)
        newdata = Mydata.active
        # print(newdata)
        for row in newdata.iter_rows(min_row=2, values_only=True):
            print(*row)
            if len(row) == 3:  # Check if the row has exactly three values
                name, age, city = row  # Extract data from each row
        # Validate data and insert into the database
                if name and age and city:
                    data = Fparser(name=name, age=age, city=city)
                    db.session.add(data)
            else:
                print("Skipping row with unexpected data:", row)
        db.session.commit()
    return "message: Data retrieved"

@app.route('/get', methods=['GET'])
@verify_token
def get_all_data():
    all_post = Fparser.query.all()
    result = parser_schemas.dump(all_post)
    return jsonify(result)

@app.route('/get/<int:id>', methods=['GET'])
@verify_token
def get_data(id):
    post = Fparser.query.filter_by(id=id).first()
    if post:
        result = parser_schema.dump(post)
        return jsonify(result)
    else:
        return jsonify({"msg": "Data not found"}), 404

@app.route('/update/<int:id>', methods=['PUT'])
@verify_token
def update_data(id):
    post = Fparser.query.get(id)
    if not post:
        return jsonify({"msg": "Data not found"}), 404

    name = request.json.get('name')
    age = request.json.get('age')
    city = request.json.get('city')

    if name:
        post.name = name
    if age:
        post.age = age
    if city:
        post.city = city

    db.session.commit()
    return jsonify({"msg": "Data updated successfully"})

@app.route('/delete/<int:id>', methods=['DELETE'])
@verify_token
def delete_data(id):
    post = Fparser.query.get(id)
    if not post:
        return jsonify({"msg": "Data not found"}), 404

    db.session.delete(post)
    db.session.commit()
    return jsonify({"msg": "Data deleted successfully"})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()

    app.run(debug=True)
